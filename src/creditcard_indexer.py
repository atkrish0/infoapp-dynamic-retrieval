from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook


CANONICAL_COLUMNS = [
    "Month",
    "Date",
    "Payee Name",
    "Type",
    "Payee / Description",
    "Category",
    "CatUpd",
    "Tag",
    "TagUpd",
    "Personal/Business",
    "P/B Only",
    "Charge",
    "Payment",
    "Credit Limit",
    "Charges",
    "Payments",
    "Available Credit",
]

HEADER_TO_FIELD = {
    "Month": "month",
    "Date": "date",
    "Payee Name": "payee_name",
    "Type": "txn_type",
    "Payee / Description": "payee_description",
    "Category": "category",
    "CatUpd": "catupd",
    "Tag": "tag",
    "TagUpd": "tagupd",
    "Personal/Business": "personal_business",
    "P/B Only": "pb_only",
    "Charge": "charge",
    "Payment": "payment",
    "Credit Limit": "credit_limit",
    "Charges": "charges",
    "Payments": "payments",
    "Available Credit": "available_credit",
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _coerce_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    # Accept existing ISO strings.
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    # Handle M/D/YYYY
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _extract_html_info_object(html_path: Path) -> tuple[str | None, list[str]]:
    html = html_path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "lxml")
    title = soup.title.text.strip() if soup.title and soup.title.text else None

    start = html.find("var infoObject=")
    mid = html.find("var infoData=")
    if start == -1 or mid == -1 or mid <= start:
        return title, []

    block = html[start + len("var infoObject=") : mid]
    cut = block.find("</script>")
    if cut != -1:
        block = block[:cut]
    block = block.strip()
    if block.endswith(";"):
        block = block[:-1].strip()

    try:
        parsed = json.loads(block)
    except json.JSONDecodeError:
        return title, []

    html_columns = [
        str(col.get("title", "")).strip()
        for col in parsed.get("orgColumns", {}).get("columns", [])
        if isinstance(col, dict)
    ]
    return title, html_columns


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS report_sources (
            report_id TEXT PRIMARY KEY,
            html_path TEXT NOT NULL,
            excel_path TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            html_title TEXT,
            html_column_count INTEGER,
            excel_column_count INTEGER,
            row_count INTEGER,
            header_match INTEGER,
            notes TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id TEXT NOT NULL,
            row_num INTEGER NOT NULL,
            month TEXT,
            date TEXT,
            payee_name TEXT,
            txn_type TEXT,
            payee_description TEXT,
            category TEXT,
            catupd TEXT,
            tag TEXT,
            tagupd TEXT,
            personal_business TEXT,
            pb_only TEXT,
            charge REAL,
            payment REAL,
            credit_limit REAL,
            charges REAL,
            payments REAL,
            available_credit REAL,
            row_json TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_transactions_report_date
        ON transactions(report_id, date);

        CREATE INDEX IF NOT EXISTS idx_transactions_report_payee
        ON transactions(report_id, payee_name);

        CREATE INDEX IF NOT EXISTS idx_transactions_report_category
        ON transactions(report_id, category);

        CREATE INDEX IF NOT EXISTS idx_transactions_report_tag
        ON transactions(report_id, tag);

        CREATE VIRTUAL TABLE IF NOT EXISTS transactions_fts USING fts5(
            content_text,
            report_id UNINDEXED,
            row_id UNINDEXED
        );
        """
    )
    conn.commit()


def _row_content_text(payload: dict[str, Any], row_num: int) -> str:
    parts = [f"row {row_num}"]
    for k, v in payload.items():
        if v is None:
            continue
        parts.append(f"{k}={v}")
    return " | ".join(parts)


def build_creditcard_index(
    excel_path: str,
    html_path: str,
    db_path: str,
    *,
    report_id: str = "creditcard",
    rebuild: bool = True,
) -> None:
    """
    Build SQLite index for a credit card report pair (HTML + Excel).
    """
    excel = Path(excel_path)
    html = Path(html_path)
    db = Path(db_path)
    db.parent.mkdir(parents=True, exist_ok=True)
    if rebuild and db.exists():
        db.unlink()

    wb = load_workbook(excel, data_only=True, read_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    header = [str(v).strip() if v is not None else "" for v in header_row]

    html_title, html_columns = _extract_html_info_object(html)

    header_match = int(header == html_columns) if html_columns else 0
    notes = []
    if html_columns and header != html_columns:
        notes.append("HTML and Excel headers differ.")
    if len(header) != len(CANONICAL_COLUMNS):
        notes.append("Excel header count differs from canonical credit-card schema.")
    notes_text = " ".join(notes) if notes else ""

    conn = sqlite3.connect(str(db))
    try:
        _ensure_schema(conn)
        conn.execute("DELETE FROM transactions WHERE report_id = ?", (report_id,))
        conn.execute("DELETE FROM transactions_fts WHERE report_id = ?", (report_id,))
        conn.execute("DELETE FROM report_sources WHERE report_id = ?", (report_id,))

        row_count = 0
        for idx, row in enumerate(rows_iter, start=1):
            if row is None or not any(v is not None and str(v).strip() != "" for v in row):
                continue

            row_map = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
            payload = {
                "month": _coerce_date(row_map.get("Month")),
                "date": _coerce_date(row_map.get("Date")),
                "payee_name": row_map.get("Payee Name"),
                "txn_type": row_map.get("Type"),
                "payee_description": row_map.get("Payee / Description"),
                "category": row_map.get("Category"),
                "catupd": row_map.get("CatUpd"),
                "tag": row_map.get("Tag"),
                "tagupd": row_map.get("TagUpd"),
                "personal_business": row_map.get("Personal/Business"),
                "pb_only": row_map.get("P/B Only"),
                "charge": _coerce_float(row_map.get("Charge")),
                "payment": _coerce_float(row_map.get("Payment")),
                "credit_limit": _coerce_float(row_map.get("Credit Limit")),
                "charges": _coerce_float(row_map.get("Charges")),
                "payments": _coerce_float(row_map.get("Payments")),
                "available_credit": _coerce_float(row_map.get("Available Credit")),
            }
            row_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
            conn.execute(
                """
                INSERT INTO transactions (
                    report_id, row_num, month, date, payee_name, txn_type, payee_description,
                    category, catupd, tag, tagupd, personal_business, pb_only, charge, payment,
                    credit_limit, charges, payments, available_credit, row_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    report_id,
                    idx,
                    payload["month"],
                    payload["date"],
                    payload["payee_name"],
                    payload["txn_type"],
                    payload["payee_description"],
                    payload["category"],
                    payload["catupd"],
                    payload["tag"],
                    payload["tagupd"],
                    payload["personal_business"],
                    payload["pb_only"],
                    payload["charge"],
                    payload["payment"],
                    payload["credit_limit"],
                    payload["charges"],
                    payload["payments"],
                    payload["available_credit"],
                    row_json,
                ),
            )
            row_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            content_text = _row_content_text(payload, idx)
            conn.execute(
                "INSERT INTO transactions_fts(content_text, report_id, row_id) VALUES (?, ?, ?)",
                (content_text, report_id, row_id),
            )
            row_count += 1

        conn.execute(
            """
            INSERT INTO report_sources (
                report_id, html_path, excel_path, sheet_name, html_title, html_column_count,
                excel_column_count, row_count, header_match, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report_id,
                str(html),
                str(excel),
                sheet_name,
                html_title,
                len(html_columns) if html_columns else None,
                len(header),
                row_count,
                header_match,
                notes_text,
                _utc_now(),
            ),
        )
        conn.commit()
    finally:
        conn.close()
